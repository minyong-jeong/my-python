import openpyxl

# ----------------------------------------------

# Write excel
wb = openpyxl.Workbook()
sheet = wb['Sheet']
sheet.title = 'MAIN'
sheet.cell(row=1, column=1, value='TEST11')
sheet.cell(row=1, column=2, value='TEST12')
sheet.cell(row=2, column=1, value='TEST21')
sheet['D1'] = 'hello'

# Create new sheet
sheet2 = wb.create_sheet('Sheet2')

# Save
wb.save('test.xlsx')

# ----------------------------------------------

# Read excel
wb = openpyxl.load_workbook("test.xlsx")
sheet = wb['MAIN']
print(sheet.cell(row=1, column=1).value)
print(sheet['D1'].value)
print("Max row: %s" % sheet.max_row)
print("Max column: %s" % sheet.max_column)

# ----------------------------------------------
